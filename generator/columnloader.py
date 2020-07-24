# -*- coding: utf-8 -*-
"""
::::::::::::::::::::::::  Critical Infrastructure Cyberspace Analysis Tool (CICAT)  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE
                                            
The contents of this material reflect the views of the author and/or the Director of the Center for Advanced Aviation 
System Development (CAASD), and do not necessarily reflect the views of the Federal Aviation Administration (FAA) 
or the Department of Transportation (DOT). Neither the FAA nor the DOT makes any warranty or guarantee, or promise, 
expressed or implied, concerning the content or accuracy of the views expressed herein. 

This is the copyright work of The MITRE Corporation and was produced for the U.S. Government under Contract Number 
DTFAWA-10-C-00080 and is subject to Federal Aviation Administration Acquisition Management System Clause 3.5-13, 
Rights in Data-General, Alt. III and Alt. IV (Oct. 1996). No other use other than that granted to the U.S. Government, 
or to those acting on behalf of the U.S. Government, under that Clause is authorized without the express written permission 
of The MITRE Corporation. For further information, please contact The MITRE Corporation, Contract Office, 7515 Colshire Drive, 
McLean, VA 22102 (703) 983-6000. ©2020 The MITRE Corporation. 

The Government retains a nonexclusive, royalty-free right to publish or reproduce this document, or to allow others to do so, for 
“Government Purposes Only.”                                           
                                            
(c) 2020 The MITRE Corporation. All Rights Reserved.

::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
columnloader.py - Imports and exports column data from/to spreadsheets
::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""

import openpyxl
from collections import defaultdict

m_DATASET = defaultdict(list )

m_testdata = [['MASTER', ['one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten']],
            ['C001', ['one', 'two', 'three', 'five', 'nine', 'eleven']],  # values not in master are not stored to or retrieved from spreadsheet
            ['C002', ['two', 'four', 'six', 'eight', 'ten']],
            ['C003', ['four', 'one', 'three', 'five']] ]

def initData():
    for t in m_testdata:
        m_DATASET[t[0]] = t[1]

"""
 Routines to store column data to a spreadsheet
 
 Spreadsheet organized to with MASTER list of values as first column.  
 MASTER is the first column in sheet containing the full list of entries
 All other columns store X's to indicate column contains MASTER entry for that row. 
 
"""

def saveMaster (sheet, dataset, ID):
    entries = dataset[ID]    
    print ('Master', ID,'contains', len(entries), 'entries')
    sheet.cell (1, 1, ID)
    i = 2   
    for k in entries:
        sheet.cell (i, 1, k)
        i = i + 1
        
def findMasterRow(dataset, val):
    indx = 1
    for j in dataset['MASTER']:
        if val == j:
            return indx
        indx = indx + 1
    return 0

def saveColumn (sheet, dataset, column, ID):
    entries = dataset[ID]    
    print ('column', ID,'contains', len(entries), 'entries')
    sheet.cell (1, column, ID)
    for k in entries:
        i = findMasterRow (dataset, k)
        if i > 0:
           sheet.cell (i+1, column, 'X')

def saveColData (wb, sheetname, dataset):
    sheet = wb.create_sheet (sheetname )    
    cols = dataset.keys()
    colindex = 1
    for c in cols:
        if colindex == 1:
            saveMaster (sheet, dataset, c)
        else:
            saveColumn (sheet, dataset, colindex, c )
        colindex = colindex + 1

# Exported interface to store dataset into spreadsheet as column data
        
def DumpColsToSpreadsheet(fname, sheetname, dataset ):
    wb = openpyxl.Workbook()   
    saveColData (wb, sheetname, dataset )
    wb.save(filename = fname )    
    
"""    
    Routines to retrieve column data from a spreadsheet

    Data retrieved into defaultdict(list) structure.  
    Column headings used as defaultdict keys.
    MASTER dictionary entry contains full list of entries
    X's from spreadsheet converted to corresponding list element from MASTER
   (Don't use X's in MASTER list...)

"""

def loadMaster (sheet ):
    ret = []
    colNum = findColbyKey (sheet, 'MASTER')
    if colNum > 0:
        for row in sheet.rows:
            val = row[colNum-1].value
            ret.append (val)
    if ret:
        del[ret[0]]
    return ret

def findColbyKey (sheet, key):   
    indx = 1
    for c in sheet.columns:
        if c[0].value == key:
            return indx
        indx = indx + 1       
    return 0
            
def loadColumn (sheet, key ):
    masterdata = loadMaster (sheet )
    ret = []
#    print ('loading column', key)
    colNum = findColbyKey (sheet, key)
    if not(colNum > 0):
        print ('column', key, 'not found.')
    else:
        rowcount = 0
        for row in sheet.rows:
            val = row[colNum-1].value
#            print('loading value', val)
            rowcount = rowcount + 1
            if not(val):
                continue 
            ret.append (masterdata[rowcount-2])

    if ret:
        del[ret[0]]
    return ret

def loadColData (wb, sheetname, dataset):
    sheet = wb[sheetname]            
    for c in sheet.columns:
#        print ('loading', c[0].value)
        dataset[c[0].value] = loadColumn (sheet, c[0].value )
           

# Call to load column data from specific spreadsheet (fname, sheetname)         
def LoadColsInSpreadsheet(fname, sheetname ):
    ret = defaultdict(list )    
    wb = openpyxl.load_workbook(fname, data_only=True)     
    loadColData (wb, sheetname, ret )
    return ret

def LoadColNames (fname, sheetname):
    ret = []
    book = openpyxl.load_workbook(fname, data_only=True)
    sheet = book[sheetname]
    for row in sheet.rows: 
        for x in row:
           ret.append(x.value)

        del ret[0]
        return ret 
    


# main entry point
if ( __name__ == "__main__"):
    
   fname = '..\data\coltest.xlsx'
 
# initialize the test data
   initData()
   
# create the spreadsheet
   DumpColsToSpreadsheet (fname, 'column test', m_DATASET)
   
# test the import  
   coldata = LoadColsInSpreadsheet (fname, 'column test' )

   print ('End of run')
   
    