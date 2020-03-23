# -*- coding: utf-8 -*-
"""
:::::::::::::::::::::::::::::::::::::  MITRE CRP PROJECT  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE

This software (or technical data) was produced for the U. S. Government under contract 355358
with Brookhaven National Laboratory, and is subject to the Rights in Data-General Clause 52.227-14 (MAY 2014) or (DEC 2007).

The following copyright notice may be affixed after receipt of written approval from the Contracting Officer.
Please contact the Contracts Management Office for assistance with obtaining approval or identifying the correct clause.
If the contract has Clause 52.227-14, Alt. IV, written approval is not required and the below copyright notice may be affixed.

(c) 2020 The MITRE Corporation. All Rights Reserved.


afactory.py - factory class for ATT&CK data

:::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""

import openpyxl
import json
import os

from amodel import TTP
from amodel import MIT
from amodel import MALWARE
from amodel import THREATACTOR
from amodel import CYBERTOOL
from amodel import ATKRELATION
from columnloader import LoadColsInSpreadsheet


def isCOA (entry):
    if not(entry.find ('course-of-action') < 0):
        return True
    return False

def isTTP(entry):
    if not(entry.find ('attack-pattern') < 0):
        return True
    return False

def isMAL(entry):
    if not(entry.find ('malware') < 0):
        return True
    return False

def isACT(entry):
    if not(entry.find('intrusion-set') < 0):
        return True
    return False

def isTOOL(entry):
    if not(entry.find('tool--') < 0):
        return True
    return False

def findOBJECT(pattern, listx):
    for t in listx:
        if (t.getID() == pattern):
            return t

def aslist (val, delimiter):
    if not(val) or val == 'None':
        return None
    elif val == '' or val == ' ' or val == 'undefined':
        return []
    elif delimiter in val: 
        return val.split(delimiter)
    else:
        return [val]

class ATTACK_FACTORY():

    # LOAD OPTIONS FOR ATTACK DATA: STIX, JSON, SPREAD, SQL
    
        def __init__ (self, loadOPT, fname, trace ):           
          if trace:
              print ('ATT&CK factory constructed..')
          self.trace = trace
          self.fname = fname
          self.loadFromSpreadsheet = False
             
          if loadOPT == 'JSON':
             if self.trace:
                 print('Opening local json dataset')
             with open(os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'data', 'ATK', 'attack.json'))) as data_file:    
               self.all_attack = json.load(data_file)

          elif loadOPT == 'STIX':
             if self.trace:
                 print ('Connecting to STIX/TAXII service')
             try:
                from attackcti import attack_client
                self.lift = attack_client()            
                self.all_attack = self.lift.get_all_stix_objects()
             except:
                print ('Cannot connect to STIX/TAXII service: ')
                print ('Most likely cause: application cannot reach the Internet')
                raise

          elif loadOPT == 'SPREAD':
              if self.trace:
                  print ('Loading data from spreadsheet', self.fname)
              self.loadFromSpreadsheet = True

          else:
              if self.trace:
                 print ('WARNING! ATTACK factory: Unsupported load option:', loadOPT)

          self.ttps = []
          self.groups = None #self.all_attack['groups']
          self.malwares = None #self.all_attack['malware']
          self.mitigations = None #self.all_attack['mitigations']
          self.techniques = None #self.all_attack['techniques']
          self.tools = None #self.all_attack['tools']  
          self.relationships = None #self.all_attack['relationships']    
          self.groupprofiles = None
               
        def loadGroupsFromJSON (self):
           ret = []
           self.groups = self.all_attack['groups']
           for x in self.groups:
                ret.append(THREATACTOR (x['created'], 
                                  x['created_by_ref'], 
                                  x['group'], 
                                  x['group_aliases'], 
                                  x['group_description'],
                                  x['group_id'], 
                                  x['group_references'], 
                                  x['id'], 
                                  x['matrix'], 
                                  x['modified'],
                                  x['type'], 
                                  x['url']   ))
                
           return ret


        def loadGroupsFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:
              ret.append(THREATACTOR (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      (row[3].value).split(),
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(),
                                      row[7].value,
                                      row[8].value,
                                      row[9].value,
                                      row[10].value,
                                      row[11].value ))
              
          del (ret[0])
          return ret

        # group profile are loaded from the ACTOR PROFILES tab in the INFRASTRUCTURE MODEL spreadsheet         
        def loadGroupProfile (self, dataset, fname, pname):
            # Profiles are created one at a time with self.groupprofiles  used as a cache for actor profile data.  
            if not (self.groupprofiles):
               self.groupprofiles =  LoadColsInSpreadsheet (fname, 'ACTOR PROFILES')
               
            if not (self.groupprofiles):
                if self.trace:
                    print ('WARNING! cannot load group profile data')
                return
                    
            ttplist = self.groupprofiles[pname]
            if not (ttplist):
                if self.trace:
                    print ('WARNING! cannot find group profile:', pname)
                return
                      
            ret = THREATACTOR(None, None, pname, pname, "", pname, None, 'intrustion-set-'+pname, 'mitre-attack', None, 'intrusion set', None )

            for j in self.groupprofiles[pname]:
                bfound = False
                for a in dataset['ATT&CK']:
                    if a.getTECHID() == j:
                        ret.addUses( a, None)
                        bfound = True
                        break

                if not(bfound):
                    for i in dataset['ATK4ICS TTPs']:
                       if i.getTECHID() == j:
                          ret.addUses( i, None)
                          bfound = True
                          break                        
               
                if not(bfound):
                    if self.trace:
                        print ('WARNING! TTP', j, 'not found')                

            return ret

        def loadGroups(self):
            if self.trace:
                print ('Loading ATT&CK Groups data..')

            if self.loadFromSpreadsheet:
                return self.loadGroupsFromSheet (self.fname, 'ATKGROUPS')
            else:
                return self.loadGroupsFromJSON ()

        def loadMalwaresFromJSON (self):
            ret = []
            self.malwares = self.all_attack['malware']
            for x in self.malwares:
                    ret.append (MALWARE (x['created'],
                                         x['created_by_ref'],
                                         x['id'],
                                         x['matrix'],
                                         x['modified'],
                                         x['software'],
                                         x['software_aliases'], 
                                         x['software_description'], 
                                         x['software_id'], 
                                         x['software_labels'], 
                                         x['software_platform'], 
                                         x['software_references'], 
                                         x['type'], 
                                         x['url']   ))
 
            return ret
        
        def loadMalwaresFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:            
                 ret.append (MALWARE (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      row[3].value,
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(','),
                                      row[7].value,
                                      row[8].value,
                                      (row[9].value).split(),
                                      (row[10].value).split(),
                                      (row[11].value).split(),
                                      row[12].value,
                                      row[13].value ))
          del (ret[0])
          return ret


        def loadMalwares(self):
            if self.trace:
                print ('Loading ATT&CK Malware data..')
                
            if self.loadFromSpreadsheet:
                return self.loadMalwaresFromSheet (self.fname, 'ATKMALWARE')
            else:
                return self.loadMalwaresFromJSON ()


        def loadMitigationsFromJSON(self):
            ret = []
            self.mitigations = self.all_attack['mitigations']
            for x in self.mitigations:
                    ret.append (MIT (x['created'],
                                           x['created_by_ref'],
                                           x['id'],
                                           x['matrix'],
                                           x['mitigation'],
                                           x['mitigation_description'],
                                           x['mitigation_references'],
                                           x['modified'],
                                           x['technique_id'],
                                           x['type'],
                                           x['url'] ))
    
            return ret

        def loadMitigationsFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:                             
              r0 = row[0].value
              r1 = row[1].value
              r2 = row[2].value
              r3 = row[3].value
              r4 = row[4].value
              r5 = row[5].value
              r6 = aslist(row[6].value, ' ')
              r7 = row[7].value
              r8 = row[8].value
              r9 = row[9].value
              r10 = row[10].value                          
              ret.append (MIT (r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10 ))
                                                 
          del (ret[0])
          return ret

        def loadMitigations(self):
            if self.trace:
                print ('Loading ATT&CK Mitigations data..')
                
            if self.loadFromSpreadsheet:
                return self.loadMitigationsFromSheet (self.fname, 'ATKMITIGATION')
            else:
                return self.loadMitigationsFromJSON ()

        def loadTechniques(self):
            if self.trace:
                print ('Loading ATT&CK TTP data..')
                
            if self.loadFromSpreadsheet:
                return self.loadTechniquesFromSheet (self.fname, 'ATT&CK')
            else:
                return self.loadTechniquesFromJSON ()

        def loadTechniquesFromJSON(self):
             ret = []
             self.techniques = self.all_attack['techniques']
             for x in self.techniques:
                     ret.append(TTP (x['capec_id'],
                                           x['capec_url'],
                                           x['contributors'],
                                           x['created'],
                                           x['created_by_ref'],
                                           x['data_sources'],
                                           x['defense_bypassed'],
                                           x['detectable_by_common_defenses'],
                                           x['detectable_explanation'],
                                           x['difficulty_explanation'],
                                           x['difficulty_for_adversary'],
                                           x['effective_permissions'],
                                           x['id'],
                                           x['matrix'],
                                           x['modified'],
                                           x['network_requirements'],
                                           x['object_marking_refs'],
                                           x['permissions_required'],
                                           x['platform'],
                                           x['remote_support'],
                                           x['system_requirements'],
                                           x['tactic'],
                                           x['tactic_type'],
                                           x['technique'],
                                           x['technique_description'],
                                           x['technique_detection'],
                                           x['technique_id'],
                                           x['technique_references'],
                                           x['type'] )) 
             self.ttps = ret
             return ret

        def loadTechniquesFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:  
              
              r0 = aslist(row[0].value, ' ')
              r1 = aslist(row[1].value, ' ')
              r2 = aslist(row[2].value, ',')
              r3 = row[3].value
              r4 = row[4].value
              r5 = aslist(row[5].value, ',')
              r6 = aslist(row[6].value, ' ')
              r7 = row[7].value
              r8 = row[8].value
              r9 = row[9].value
              r10 = row[10].value
              r11 = aslist(row[11].value, ' ')
              r12 = row[12].value
              r13 = row[13].value
              r14 = row[14].value              
              r15 = row[15].value               
              r16 = aslist(row[16].value, ' ')              
              r17 = aslist(row[17].value, ' ')               
              r18 = aslist(row[18].value, ' ') 
              r19 = row[19].value 
              r20 = aslist(row[20].value, ' ')
              r21 = aslist(row[21].value, ' ') 
              r22 = row[22].value 
              r23 = row[23].value 
              r24 = row[24].value 
              r25 = row[25].value 
              r26 = row[26].value 
              r27 = aslist(row[27].value, ' ') 
              r28 = row[28].value               
                           
              ret.append (TTP (r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10,
                               r11, r12, r13, r14, r15, r16, r17, r18, r19,
                               r20, r21, r22, r23, r24, r25, r26, r27, r28 ) )
         
          del (ret[0])
          return ret             
             

        def loadTools(self):
            if self.trace:
                print ('Loading ATT&CK Tools data..')
            if self.loadFromSpreadsheet:
                return self.loadToolsFromSheet (self.fname, 'ATKTOOL' )
            else:
                return self.loadToolsFromJSON ( )
            
        def loadToolsFromJSON(self):
            ret = []
            self.tools = self.all_attack['tools']
            for x in self.tools:
                    ret.append(CYBERTOOL (x['created'],
                                    x['created_by_ref'],
                                    x['id'],
                                    x['matrix'],
                                    x['modified'],
                                    x['software'],
                                    x['software_aliases'],
                                    x['software_description'],
                                    x['software_id'],
                                    x['software_labels'],
                                    x['software_platform'],
                                    x['software_references'],
                                    x['type'], 
                                    x['url']  ))            
            return ret             


        def loadToolsFromSheet(self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows: 
                    ret.append(CYBERTOOL (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      row[3].value,
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(','),
                                      row[7].value,
                                      row[8].value,
                                      (row[9].value).split(),
                                      (row[10].value).split(),
                                      (row[11].value).split(),
                                      row[12].value,
                                      row[13].value ))     
          del (ret[0])
          return ret             


        def findTTP(self, ID, dataset):
#            if not(self.ttps):
#                print ('No TTPs to find.' )
#                return            
            for j in dataset['ATT&CK']:
                if (ID == j.getTECHID()):
                    return j                        


        def loadTTPExtension (self, dataset, filename, sheetname ):   
          if self.trace:
              print ('Loading ATT&CK extensions:', sheetname)
          book = openpyxl.load_workbook(filename, data_only=True) 
          sheet = book[sheetname]       
          for row in sheet.rows:
            ttp = self.findTTP(row[0].value, dataset)
            if (ttp):
                ttp.setP(row[1].value )
                       

        def find (self, pat, myDATAWARE ):
            if isCOA(pat):
               return findOBJECT(pat, myDATAWARE['ATKMITIGATION'] )
            elif isTTP(pat):
               return findOBJECT(pat, myDATAWARE['ATT&CK'])
            elif isACT(pat):
               return findOBJECT(pat, myDATAWARE['ATKGROUPS'])
            elif isMAL(pat):
               return findOBJECT (pat, myDATAWARE['ATKMALWARE'])
            elif isTOOL(pat):
               return findOBJECT(pat, myDATAWARE['ATKTOOL'])
           

        def initRelationships(self, data):
            if self.trace:
                print ('Loading ATT&CK Relationships data..')
            if self.loadFromSpreadsheet:
                return self.loadRelationshipsFromSheet (self.fname, 'ATKRELS', data)
            else:
                return self.loadRelationshipsFromJSON ( data )

        def loadRelationshipsFromJSON(self, data ):
            
            ret = []
            countr=0
            self.relationships = self.all_attack['relationships']
            for rel in self.relationships:               
               ret.append(ATKRELATION (rel['created'],
                                       rel['created_by_ref'],
                                       rel['id'],
                                       rel['modified'],
                                       rel['relationship'],
                                       rel['relationship_description'], 
                                       rel['source_object'], 
                                       rel['target_object'] ))               
                                
               countr = countr+1
               src = self.find (rel['source_object'], data)
               if not(src):
                   if self.trace:
                      print ('loadRelationships(): Relation', countr, ': SRC', rel['source_object'], 'not found.')
       
               tgt = self.find (rel['target_object'], data)
               if not(tgt):
                   if self.trace:
                      print ('loadRelationships(): Relation', countr, ': TGT', rel['target_object'], 'not found.')
        
               if (src) and (tgt):
                  desc = rel['relationship_description']
                  if (rel['relationship']=='uses'):
                    src.addUses(tgt, desc)
                  elif (rel['relationship']=='mitigates'):
                    src.addMitigates(tgt, desc)
                    tgt.addCOA(src, desc)
                  elif (rel['relationship']=='relates-to'):
                    src.addRelates (tgt, desc)
                  elif (rel['relationship']=='revoked-by'):
                    src.addRevokes(tgt, desc)
         
            return ret
            
        def loadRelationshipsFromSheet(self, fname, sname, data ):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]

          ret = []         
          countr=0
          for rel in sheet.rows:
              
               ret.append(ATKRELATION (rel[0].value,
                                       rel[1].value,
                                       rel[2].value,
                                       rel[3].value,
                                       rel[4].value,
                                       rel[5].value, 
                                       rel[6].value, 
                                       rel[7].value ))       
              
               countr = countr+1
               src = self.find (rel[6].value, data)
       
               tgt = self.find (rel[7].value, data)
        
               if (src) and (tgt):
                  desc = rel[5].value
                  if (rel[4].value =='uses'):
                    src.addUses(tgt, desc)
                  elif (rel[4].value =='mitigates'):
                    src.addMitigates(tgt, desc)
                    tgt.addCOA(src, desc)
                  elif (rel[4].value =='relates-to'):
                    src.addRelates (tgt, desc)
                  elif (rel[4].value =='revoked-by'):
                    src.addRevokes(tgt, desc)

          del (ret[0])
          return ret
     

        def exportRelationships(self):
            ret = []
            if self.relationships:
              for rel in self.relationships:
                ret.append(ATKRELATION (rel['created'],
                                        rel['created_by_ref'],
                                        rel['id'],
                                        rel['modified'],
                                        rel['relationship'],
                                        rel['relationship_description'], 
                                        rel['source_object'], 
                                        rel['target_object'] ))
            
            return ret
   
