# -*- coding: utf-8 -*-
"""
::::::::::::::::::::::::  Critical Infrastructure Cyberspace Analysis Tool (CICAT)  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE
                                            
The contents of this material reflect the views of the author and/or the Director of the Center for Advanced Aviation 
System Development (CAASD), and do not necessarily reflect the views of the Federal Aviation Administration (FAA) 
or the Department of Transportation (DOT). Neither the FAA nor the DOT makes any warranty or guarantee, or promise, 
expressed or implied, concerning the content or accuracy of the views expressed herein. 

This is the copyright work of The MITRE Corporation and was produced for the U.S. Government under Contract Number 
DTFAWA-10-C-00080 and is subject to Federal Aviation Administration Acquisition Management System Clause 3.5-13, 
Rights in Data-General, Alt. III and Alt. IV (Oct. 1996). No other use other than that granted to the U.S. Government, 
or to those acting on behalf of the U.S. Government, under that Clause is authorized without the express written permission 
of The MITRE Corporation. For further information, please contact The MITRE Corporation, Contract Office, 7515 Colshire Drive, 
McLean, VA 22102 (703) 983-6000. ©2020 The MITRE Corporation. 

The Government retains a nonexclusive, royalty-free right to publish or reproduce this document, or to allow others to do so, for 
“Government Purposes Only.”                                           
                                            
(c) 2020 The MITRE Corporation. All Rights Reserved.

::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
tfactory.py - Factory class to construct target, scenario, entrypoint object instances
::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""


import openpyxl

from tmodel import TARGET
from tmodel import SCENARIO
#from tmodel import INDICATOR
#from tmodel import COA
from tmodel import ENTRYPOINT

def findActor(aid, dataset ):
    for j in dataset['ATKGROUPS']:
        if j.getGroupID() == aid:
            return j
        
def findTargetRecord(tid, dataset):
    for j in dataset['TARGET']:
        if j.getName() == tid:
            return j
        
def findTarget(tgtobj, dataset ):
    tname = tgtobj.getName()    
    for j in dataset['COMPONENT']:
        if j.getName().lower() == tname.lower():
            return j


class THREAT_FACTORY():
    def __init__ (self, filename, trace ):
        self.filename = filename
        self.trace = trace
        if self.trace:
            print ('THREAT factory constructed..')
        
    def getLoader (self, sheetname ): 
        if (sheetname == 'TARGET'):
            return INFRASTRUCTURE_FACTORY (self.filename, 'INFRASTRUCTURE', True, self.trace )
        elif (sheetname == 'ENTRYPOINT'):
            return INFRASTRUCTURE_FACTORY (self.filename, 'INFRASTRUCTURE', False, self.trace )
        elif (sheetname == 'SCENARIO'):
            return SCENARIO_FACTORY(self.filename, sheetname, self.trace)
#        elif (sheetname == 'COA'):
#            return COA_FACTORY (self.filename, sheetname, self.trace)
        return
    
    def initRelationships(self, dataset ):       
      for s in dataset['SCENARIO']:
        tid = s.getTargetID()
        target = findTargetRecord(tid, dataset)
        if target:
             s.setTarget(target )
        else:
            if self.trace:
               print ('WARNING: Could not find target', tid)
            
        aid = s.getActorID()
        actor = findActor (aid, dataset)
        if actor:
            s.setActor(actor)
        else:
            if self.trace:
               print ('WARNING: Could not find actor', aid)

"""
class INDICATOR_FACTORY(THREAT_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('INDICATOR factory constructed..')
        return
    
    def load (self):
       if self.trace:
           print ('Loading Indicators..')
       book = openpyxl.load_workbook(self.filename, data_only=True) 
       sheet = book[self.sheetname]
       ret = []         
       for row in sheet.rows:
         ret.append(INDICATOR (row[0].value, row[1].value  ))          

       del ret[0]
       return ret
"""
     
class INFRASTRUCTURE_FACTORY (THREAT_FACTORY):
    def __init__ (self, filename, sheetname, btargetFlag, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.btargetFlag = btargetFlag
        self.trace = trace
        if self.trace:
            if btargetFlag:
               print ('TARGET factory constructed..')
            else:
               print ('ENTRYPOINT factory constructed..')      
        return

    def load (self):
        if self.trace:
               if self.btargetFlag:
                   print ('Loading target information..')
               else:
                   print ('Loading entry point information..')
        book = openpyxl.load_workbook(self.filename, data_only=True) 
        sheet = book[self.sheetname]
        ret = []
        
        if self.btargetFlag:
           for row in sheet.rows:
               if row[3].value:
                 ret.append (TARGET (row[0].value, row[1].value ) )        
        else:
            for row in sheet.rows:
                if row[4].value:
                 ret.append (ENTRYPOINT (row[0].value, row[1].value ))
        del ret[0]
        return ret   

"""
class COA_FACTORY (THREAT_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('COA factory constructed..')
        return

    def load (self, filename):
        if self.trace:
            print ('Loading COAs..')
        book = openpyxl.load_workbook(filename, data_only=True) 
        sheet = book['Controls']
        ret = []
        control = None
        for row in sheet.rows:
            if (row[5].value.startswith ('[Withdrawn:')):
                continue
            if row[0].value and row[2].value:
                control = COA (row[0].value, row[1].value, row[2].value, row[3].value, 
                               row[4].value, row[5].value, row[6].value, row[7].value )
                ret.append(control)
            else:
                control.appendDesc(row[1].value, row[5].value)

        del ret[0]
        return ret       
"""

class SCENARIO_FACTORY(THREAT_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('SCENARIO factory constructed..')
        return
    
    def load (self):
        if self.trace:
            print ('Loading Scenarios..')
          
        book = openpyxl.load_workbook(self.filename, data_only=True) 
        sheet = book[self.sheetname]
        ret = []
        for row in sheet.rows:
           ret.append (SCENARIO (row[0].value, row[1].value, row[2].value, 
                                  row[3].value, row[4].value, row[5].value, 
                                  row[6].value, row[7].value ) )

        del ret[0]
        return ret
    
