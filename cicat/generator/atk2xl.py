# -*- coding: utf-8 -*-
"""
::::::::::::::::::::::::  Critical Infrastructure Cyberspace Analysis Tool (CICAT)  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE
                                            
The contents of this material reflect the views of the author and/or the Director of the Center for Advanced Aviation 
System Development (CAASD), and do not necessarily reflect the views of the Federal Aviation Administration (FAA) 
or the Department of Transportation (DOT). Neither the FAA nor the DOT makes any warranty or guarantee, or promise, 
expressed or implied, concerning the content or accuracy of the views expressed herein. 

This is the copyright work of The MITRE Corporation and was produced for the U.S. Government under Contract Number 
DTFAWA-10-C-00080 and is subject to Federal Aviation Administration Acquisition Management System Clause 3.5-13, 
Rights in Data-General, Alt. III and Alt. IV (Oct. 1996). No other use other than that granted to the U.S. Government, 
or to those acting on behalf of the U.S. Government, under that Clause is authorized without the express written permission 
of The MITRE Corporation. For further information, please contact The MITRE Corporation, Contract Office, 7515 Colshire Drive, 
McLean, VA 22102 (703) 983-6000. ©2020 The MITRE Corporation. 

The Government retains a nonexclusive, royalty-free right to publish or reproduce this document, or to allow others to do so, for 
“Government Purposes Only.”                                           
                                            
(c) 2020 The MITRE Corporation. All Rights Reserved.

::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
atk2xl.py - Utility to load ATTACK data into spreadsheet
::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""

import datetime
import sys
from openpyxl import Workbook
from collections import defaultdict
from loaddata import LOAD_DATA
from loaddata import m_file_INFRASTRUCTURE, m_file_SCENARIOS 

m_DATASET = defaultdict(list )

m_ATTCKSHEETS = ['ATT&CK', 'ATKMITIGATION', 'ATKGROUPS', 'ATKMALWARE','ATKTOOL', 'ATKRELS'] 

def list2string(plist, separator):
    ret = 'undefined'
    if plist:
       ret = str(plist[0])
       for j in range (1, len(plist)):
           ret = ret + separator + str(plist[j])
    return ret

def saveTechHeader(sheet):
    sheet.cell(1, 1, 'capec_id')
    sheet.cell(1, 2, 'capec_url')
    sheet.cell(1, 3, 'contributors')
    sheet.cell(1, 4, 'date_created')
    sheet.cell(1, 5, 'created_by_ref')
    sheet.cell(1, 6, 'data_sources')
    sheet.cell(1, 7, 'defense_bypassed')
    sheet.cell(1, 8, 'detectable_by_common_defenses')
    sheet.cell(1, 9, 'detectable_explanation')      
    sheet.cell(1, 10, 'difficulty_explanation')
    sheet.cell(1, 11, 'difficulty_for_adversary)')
    sheet.cell(1, 12, 'effective_permissions')
    sheet.cell(1, 13, 'ID')
    sheet.cell(1, 14, 'matrix')
    sheet.cell(1, 15, 'date_modified')
    sheet.cell(1, 16, 'network_requirements')
    sheet.cell(1, 17, 'object_marking_refs')
    sheet.cell(1, 18, 'permissions_required')
    sheet.cell(1, 19, 'platform')
    sheet.cell(1, 20, 'remote_support')
    sheet.cell(1, 21, 'system_requirements')
    sheet.cell(1, 22, 'tactic')
    sheet.cell(1, 23, 'tactic_type')
    sheet.cell(1, 24, 'tech_name')
    sheet.cell(1, 25, 'tech_desc')
    sheet.cell(1, 26, 'tech_detect')
    sheet.cell(1, 27, 'tech_id')
    sheet.cell(1, 28, 'tech_references')
    sheet.cell(1, 29, 'type' )    

def saveTechnique(sheet, row, TTP):    
    sheet.cell(row, 1, list2string(TTP.capec_id))
    sheet.cell(row, 2, list2string(TTP.capec_url))
    sheet.cell(row, 3, list2string(TTP.contributors))
    sheet.cell(row, 4, str(TTP.created))
    sheet.cell(row, 5, str(TTP.created_by_ref))
    sheet.cell(row, 6, list2string(TTP.data_sources))
    sheet.cell(row, 7, list2string(TTP.defense_bypassed))
    sheet.cell(row, 8, str(TTP.detectable_by_common_defenses ))
    sheet.cell(row, 9, str(TTP.detectable_explanation))      
    sheet.cell(row, 10, str(TTP.difficulty_explanation))
    sheet.cell(row, 11, str(TTP.difficulty_for_adversary))
    sheet.cell(row, 12, list2string(TTP.effective_permissions))
    sheet.cell(row, 13, str(TTP.myID))
    sheet.cell(row, 14, str(TTP.matrix))
    sheet.cell(row, 15, str(TTP.modified))
    sheet.cell(row, 16, str(TTP.network_requirements))
    sheet.cell(row, 17, list2string(TTP.object_marking_refs))
    sheet.cell(row, 18, list2string(TTP.permissions_required))
    sheet.cell(row, 19, list2string(TTP.platform ))
    sheet.cell(row, 20, str(TTP.remote_support))
    sheet.cell(row, 21, list2string(TTP.system_requirements))
    sheet.cell(row, 22, list2string(TTP.tactic ))
    sheet.cell(row, 23, str(TTP.tactic_type))
    sheet.cell(row, 24, str(TTP.technique))
    sheet.cell(row, 25, str(TTP.desc))
    sheet.cell(row, 26, str(TTP.technique_detection))
    sheet.cell(row, 27, str(TTP.technique_id))
    sheet.cell(row, 28, list2string(TTP.technique_references))
    sheet.cell(row, 29, str(TTP.typex))    

def saveICSTTP (sheet, row, ICSTTP):
    sheet.cell(row, 1, 'undefined' ) #list2string(TTP.capec_id))
    sheet.cell(row, 2, 'undefined' ) #list2string(TTP.capec_url))
    sheet.cell(row, 3, ICSTTP.contributor )
    sheet.cell(row, 4, datetime.datetime.now() )
    sheet.cell(row, 5, 'identity--c78cb6e5-0c4b-4611-8297-d1b8b55e40b5' ) #str(TTP.created_by_ref))
    sheet.cell(row, 6, list2string(ICSTTP.datasources, ','))
    sheet.cell(row, 7, 'undefined' ) #list2string(TTP.defense_bypassed))
    sheet.cell(row, 8, None ) #str(TTP.detectable_by_common_defenses ))
    sheet.cell(row, 9, None ) #str(TTP.detectable_explanation))      
    sheet.cell(row, 10, None) #str(TTP.difficulty_explanation))
    sheet.cell(row, 11, None) #str(TTP.difficulty_for_adversary))
    sheet.cell(row, 12, 'undefined' ) #list2string(TTP.effective_permissions))
    sheet.cell(row, 13, 'attack4ics-pattern--'+str(ICSTTP.aid))
    sheet.cell(row, 14, 'mitre-attack4ics')
    sheet.cell(row, 15, None ) #str(TTP.modified))
    sheet.cell(row, 16, None ) #str(TTP.network_requirements))
    sheet.cell(row, 17, None ) #list2string(TTP.object_marking_refs))
    sheet.cell(row, 18, None) #list2string(TTP.permissions_required))
    sheet.cell(row, 19, list2string(ICSTTP.platform, ',' ))
    sheet.cell(row, 20, None ) #str(TTP.remote_support))
    sheet.cell(row, 21, 'undefined' ) #list2string(TTP.system_requirements))
    sheet.cell(row, 22, list2string(ICSTTP.tactic, ',' ))
    sheet.cell(row, 23, None ) #str(TTP.tactic_type))
    sheet.cell(row, 24, str(ICSTTP.name))
    sheet.cell(row, 25, str(ICSTTP.desc))
    sheet.cell(row, 26, None ) #str(TTP.technique_detection))
    sheet.cell(row, 27, str(ICSTTP.techID))
    sheet.cell(row, 28, ICSTTP.url)
    sheet.cell(row, 29, 'attack-pattern')    


def saveMitHeader(sheet, row=1 ): 
    sheet.cell(row, 1, 'date_created')
    sheet.cell(row, 2, 'created_by_ref')
    sheet.cell(row, 3, 'ID')
    sheet.cell(row, 4, 'matrix')
    sheet.cell(row, 5, 'mitigation')
    sheet.cell(row, 6, 'mit_desc')
    sheet.cell(row, 7, 'mit_references')
    sheet.cell(row, 8, 'date_modified')
    sheet.cell(row, 9, 'mit_technique_id')
    sheet.cell(row, 10, 'type')
    sheet.cell(row, 11, 'mit_url')

def saveICSMitigation(sheet, row, ICSTTP): 
    sheet.cell(row, 1, datetime.datetime.now() )
    sheet.cell(row, 2, 'identity--c78cb6e5-0c4b-4611-8297-d1b8b55e40b5')  #str(MIT.created_by_ref))
    sheet.cell(row, 3, 'attack4ics-course-of-action--'+str(ICSTTP.aid) )
    sheet.cell(row, 4, 'mitre-attack4ics' )
    sheet.cell(row, 5, ICSTTP.name + ' Mitigation')
    sheet.cell(row, 6, ICSTTP.mitigation)
    sheet.cell(row, 7, ICSTTP.url)
    sheet.cell(row, 8, None ) #str(MIT.modified))
    sheet.cell(row, 9, ICSTTP.techID )
    sheet.cell(row, 10, 'course-of-action')
    sheet.cell(row, 11, ICSTTP.url )
    
def saveMitigation(sheet, row, MIT ): 
    sheet.cell(row, 1, str(MIT.created))
    sheet.cell(row, 2, str(MIT.created_by_ref))
    sheet.cell(row, 3, str(MIT.myID))
    sheet.cell(row, 4, str(MIT.matrix))
    sheet.cell(row, 5, str(MIT.mitigation))
    sheet.cell(row, 6, str(MIT.desc))
    sheet.cell(row, 7, list2string(MIT.mitigation_references))
    sheet.cell(row, 8, str(MIT.modified))
    sheet.cell(row, 9, str(MIT.technique_id))
    sheet.cell(row, 10, str(MIT.typex))
    sheet.cell(row, 11, str(MIT.url))
    
def saveMalHeader(sheet, row=1):
    sheet.cell(row, 1, 'date_created')
    sheet.cell(row, 2, 'created_by_ref')
    sheet.cell(row, 3, 'ID')
    sheet.cell(row, 4, 'matrix')
    sheet.cell(row, 5, 'date_modified')
    sheet.cell(row, 6, 'software')
    sheet.cell(row, 7, 'software_aliases')
    sheet.cell(row, 8, 'desc')
    sheet.cell(row, 9, 'software_id')
    sheet.cell(row, 10, 'software_labels')
    sheet.cell(row, 11, 'software_platform' )
    sheet.cell(row, 12, 'software_references')
    sheet.cell(row, 13, 'typex')
    sheet.cell(row, 14, 'mal_url')

def saveMalware(sheet, row, MAL):
    sheet.cell(row, 1, str(MAL.created ))
    sheet.cell(row, 2, str(MAL.created_by_ref))
    sheet.cell(row, 3, str(MAL.myID ))
    sheet.cell(row, 4, str(MAL.matrix ))
    sheet.cell(row, 5, str(MAL.modified ))
    sheet.cell(row, 6, str(MAL.software ))
    sheet.cell(row, 7, list2string(MAL.software_aliases ))
    sheet.cell(row, 8, str(MAL.desc ))
    sheet.cell(row, 9, str(MAL.software_id ))
    sheet.cell(row, 10, list2string(MAL.software_labels ))
    sheet.cell(row, 11, list2string(MAL.software_platform ))
    sheet.cell(row, 12, list2string(MAL.software_references ))
    sheet.cell(row, 13, str(MAL.typex ))
    sheet.cell(row, 14, str(MAL.url ))

def saveToolHeader(sheet, row=1):
    sheet.cell(row, 1, 'date_created')
    sheet.cell(row, 2, 'created_by_ref')
    sheet.cell(row, 3, 'ID')
    sheet.cell(row, 4, 'matrix')
    sheet.cell(row, 5, 'date_modified')
    sheet.cell(row, 6, 'software')
    sheet.cell(row, 7, 'software_aliases')
    sheet.cell(row, 8, 'desc')
    sheet.cell(row, 9, 'software_id')
    sheet.cell(row, 10, 'software_labels')
    sheet.cell(row, 11, 'software_platform')
    sheet.cell(row, 12, 'software_references')
    sheet.cell(row, 13, 'type')
    sheet.cell(row, 14, 'tool_url')    
    
def saveCybertool(sheet, row, TOOL):
    sheet.cell(row, 1, str(TOOL.created ))
    sheet.cell(row, 2, str(TOOL.created_by_ref ))
    sheet.cell(row, 3, str(TOOL.myID ))
    sheet.cell(row, 4, str(TOOL.matrix ))
    sheet.cell(row, 5, str(TOOL.modified ))
    sheet.cell(row, 6, str(TOOL.software ))
    sheet.cell(row, 7, list2string(TOOL.software_aliases ))
    sheet.cell(row, 8, str(TOOL.desc ))
    sheet.cell(row, 9, str(TOOL.software_id ))
    sheet.cell(row, 10, list2string(TOOL.software_labels ))
    sheet.cell(row, 11, list2string(TOOL.software_platform ))
    sheet.cell(row, 12, list2string(TOOL.software_references ))
    sheet.cell(row, 13, str(TOOL.typex ))
    sheet.cell(row, 14, str(TOOL.url ))    
    
def saveGroupHeader(sheet, row=1):   
    sheet.cell(row, 1, 'date_created')
    sheet.cell(row, 2, 'created_by_ref')
    sheet.cell(row, 3, 'group')
    sheet.cell(row, 4, 'group_aliases')
    sheet.cell(row, 5, 'desc')
    sheet.cell(row, 6, 'group_id')
    sheet.cell(row, 7, 'group_references')
    sheet.cell(row, 8, 'ID')
    sheet.cell(row, 9, 'matrix')
    sheet.cell(row, 10, 'date_modified')
    sheet.cell(row, 11, 'type')
    sheet.cell(row, 12, 'group_url')   

def saveGroup(sheet, row, ACT):   
    sheet.cell(row, 1, str(ACT.created ))
    sheet.cell(row, 2, str(ACT.created_by_ref ))
    sheet.cell(row, 3, str(ACT.group ))
    sheet.cell(row, 4, list2string(ACT.group_aliases ))
    sheet.cell(row, 5, str(ACT.desc ))
    sheet.cell(row, 6, str(ACT.group_id ))
    sheet.cell(row, 7, list2string(ACT.group_references ))
    sheet.cell(row, 8, str(ACT.myID ))
    sheet.cell(row, 9, str(ACT.matrix ))
    sheet.cell(row, 10, str(ACT.modified ))
    sheet.cell(row, 11, str(ACT.typex ))
    sheet.cell(row, 12, str(ACT.url ))   
    
def saveRelationHeader(sheet, row=1):   
    sheet.cell(row, 1, 'date_created')
    sheet.cell(row, 2, 'created_by_ref')
    sheet.cell(row, 3, 'ID')
    sheet.cell(row, 4, 'date_modified')
    sheet.cell(row, 5, 'rship_name')
    sheet.cell(row, 6, 'rship_desc')
    sheet.cell(row, 7, 'src_obj')
    sheet.cell(row, 8, 'tgt_obj')
        
def saveRelation(sheet, row, REL):   
    sheet.cell(row, 1, str(REL.created ))
    sheet.cell(row, 2, str(REL.created_by_ref ))
    sheet.cell(row, 3, str(REL.myID ))
    sheet.cell(row, 4, str(REL.modified ))
    sheet.cell(row, 5, str(REL.rship_name ))
    sheet.cell(row, 6, str(REL.rship_desc ))
    sheet.cell(row, 7, str(REL.src_obj ))
    sheet.cell(row, 8, str(REL.tgt_obj ))
        
def saveDataset (wb, sheetname):

    sheet = wb.create_sheet (title=sheetname )
    if sheetname == 'ATT&CK':
        saveTechHeader (sheet)
        i=2
        for row in m_DATASET[sheetname]:
            saveTechnique(sheet, i, row)
            i = i + 1
                        
    elif sheetname == 'ATKMITIGATION':
        saveMitHeader(sheet)
        i = 2
        for row in m_DATASET[sheetname]:
           saveMitigation(sheet, i, row)
           i= i + 1

    elif sheetname == 'ATKGROUPS':
        saveGroupHeader(sheet)
        i = 2
        for row in m_DATASET[sheetname]:    
           saveGroup(sheet, i, row)
           i= i + 1

    elif sheetname == 'ATKMALWARE':
         saveMalHeader(sheet)
         i = 2
         for row in m_DATASET[sheetname]:
            saveMalware(sheet, i, row)
            i= i + 1
            
    elif sheetname == 'ATKTOOL':
         saveToolHeader(sheet)
         i = 2
         for row in m_DATASET[sheetname]:
            saveCybertool(sheet, i, row)
            i= i + 1

    elif sheetname == 'ATKRELS':
         saveRelationHeader(sheet)
         i = 2
         for row in m_DATASET[sheetname]:
            saveRelation(sheet, i, row)
            i= i + 1
    else:
         print ('Dont know', sheetname)
         
def saveICSDataset (wb, sheetname, dataset):
    
    sheet = wb.create_sheet(title=sheetname )
    if sheetname == 'ATK4ICS TTPs':
         saveTechHeader (sheet)        
         i = 2
         for row in dataset.keys():
             saveICSTTP (sheet, i, dataset[row] )
             i = i + 1

    elif sheetname == 'ATK4ICS MITs':
        saveMitHeader(sheet)
        i = 2
        for row in dataset.keys():
           saveICSMitigation(sheet, i, dataset[row])
           i = i + 1
        

def DumpToSpreadsheet(fname ):
    wb = Workbook()     
    for sheet in m_ATTCKSHEETS:
        saveDataset (wb, sheet)
    wb.save(filename = fname )    


# main entry point
if ( __name__ == "__main__"):
    
#  default settings
   fname = '../data/ATTACK.xlsx'
   loadopt = 'JSON'
   
   params = sys.argv
   if (len(params) > 1 ): 
       fname = params[1]    
       
   if (len(params) > 2 ):
       loadopt = params[2]
       
   if not(loadopt == 'JSON') and not(loadopt == 'STIX'):
       print ('Usage: python', params[0], '<filename.xlsx> [JSON|STIX]' )
       exit()
   
   print ('Loading ATT&CK data from', loadopt, 'into spreadsheet', fname)
   
   m_DATASET = LOAD_DATA (m_file_INFRASTRUCTURE, m_file_SCENARIOS, False, False)   
   DumpToSpreadsheet (fname )

   print ('End of run')
